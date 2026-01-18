from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils.translation import gettext as _
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import SessionAuthentication, TokenAuthentication
from core.authentication import BearerTokenAuthentication
from rest_framework.response import Response
from .models import Campaign, Email, EmailEvent, EmailAIAnalysis
from django.urls import reverse
from django.template import TemplateDoesNotExist
from .serializers import CampaignSerializer, EmailSerializer
from subscribers.models import List
import csv
import io
import logging
from openpyxl import load_workbook
from .ai_utils import generate_email_content

logger = logging.getLogger(__name__)

@login_required
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def campaign_list_create(request):
    """List all campaigns or create a new one."""
    if request.method == 'GET':
        campaigns = Campaign.objects.filter(user=request.user)
        serializer = CampaignSerializer(campaigns, many=True)
        return Response(serializer.data)
    
    serializer = CampaignSerializer(data=request.data, context={'request': request})
    if serializer.is_valid():
        campaign = serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)

@login_required
@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def campaign_detail(request, campaign_id):
    """Get, update or delete a campaign."""
    campaign = get_object_or_404(Campaign, id=campaign_id, user=request.user)
    
    if request.method == 'GET':
        serializer = CampaignSerializer(campaign)
        return Response(serializer.data)
    
    elif request.method == 'PUT':
        serializer = CampaignSerializer(campaign, data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)
    
    elif request.method == 'DELETE':
        campaign.delete()
        return Response(status=204)

@login_required
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def email_list_create(request, campaign_id):
    """List all emails in a campaign or create a new one."""
    from django.utils import timezone
    from datetime import timedelta
    from campaigns.models import EmailSendRequest
    from subscribers.models import Subscriber
    import logging
    
    logger = logging.getLogger(__name__)
    campaign = get_object_or_404(Campaign, id=campaign_id, user=request.user)
    
    if request.method == 'GET':
        emails = campaign.emails.all().order_by('order')
        serializer = EmailSerializer(emails, many=True)
        return Response(serializer.data)
    
    elif request.method == 'POST':
        # Set the order to be the next available number
        data = request.data.copy()
        data['order'] = campaign.emails.count()
        # Remove data['campaign'] = campaign.id
        serializer = EmailSerializer(data=data, context={'request': request})
        if serializer.is_valid():
            email = serializer.save(campaign=campaign)
            
            # Check if campaign has already sent any emails
            has_sent_emails = EmailSendRequest.objects.filter(
                campaign=campaign,
                status='sent'
            ).exists()
            
            # If campaign has sent emails, automatically schedule this new email to all active subscribers
            if has_sent_emails and campaign.subscriber_list:
                # Get all active subscribers from the campaign's subscriber list
                active_subscribers = Subscriber.objects.filter(
                    lists=campaign.subscriber_list,
                    is_active=True
                ).distinct()
                
                if active_subscribers.exists():
                    # Calculate scheduled_for time based on email's wait_time and wait_unit
                    # Use 0 as default if wait_time is None, but allow 0 to mean immediate sending
                    wait_time = email.wait_time if email.wait_time is not None else 0
                    wait_unit = email.wait_unit or 'days'
                    
                    send_delay = timedelta(0)
                    if wait_time > 0:
                        if wait_unit == 'seconds':
                            send_delay = timedelta(seconds=wait_time)
                        elif wait_unit == 'minutes':
                            send_delay = timedelta(minutes=wait_time)
                        elif wait_unit == 'hours':
                            send_delay = timedelta(hours=wait_time)
                        elif wait_unit == 'days':
                            send_delay = timedelta(days=wait_time)
                        elif wait_unit == 'weeks':
                            send_delay = timedelta(weeks=wait_time)
                        elif wait_unit == 'months':
                            send_delay = timedelta(days=wait_time * 30)
                    
                    scheduled_for = timezone.now() + send_delay
                    
                    # Create EmailSendRequest for each active subscriber
                    send_requests = []
                    for subscriber in active_subscribers:
                        variables = {
                            'first_name': subscriber.first_name or '',
                            'last_name': subscriber.last_name or '',
                            'email': subscriber.email
                        }
                        
                        send_request = EmailSendRequest.objects.create(
                            user=request.user,
                            campaign=campaign,
                            email=email,
                            subscriber=subscriber,
                            subscriber_email=subscriber.email,
                            variables=variables,
                            scheduled_for=scheduled_for,
                            status='pending'
                        )
                        send_requests.append(send_request)
                    
                    logger.info(f"Auto-scheduled email '{email.subject}' to {len(send_requests)} active subscribers in campaign '{campaign.name}'")
            
            return Response(serializer.data, status=201)
        return Response({'error': 'Validation failed', 'details': serializer.errors}, status=400)

@login_required
@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def email_detail(request, campaign_id, email_id):
    """Get, update or delete an email."""
    campaign = get_object_or_404(Campaign, id=campaign_id, user=request.user)
    email = get_object_or_404(Email, id=email_id, campaign=campaign)
    
    if request.method == 'GET':
        serializer = EmailSerializer(email)
        return Response(serializer.data)
    
    elif request.method == 'PUT':
        from .models import EmailSendRequest
        from subscribers.models import Subscriber
        from django.utils import timezone
        from datetime import timedelta
        import logging
        
        logger = logging.getLogger(__name__)
        
        serializer = EmailSerializer(email, data=request.data, context={'request': request})
        if serializer.is_valid():
            email = serializer.save()
            
            # Check if campaign has already sent any emails
            has_sent_emails = EmailSendRequest.objects.filter(
                campaign=campaign,
                status='sent'
            ).exists()
            
            # Check if this email has already been sent or scheduled to all active subscribers
            # If not, and campaign has sent emails, schedule it to all active subscribers
            if has_sent_emails and campaign.subscriber_list:
                # Get all active subscribers from the campaign's subscriber list
                active_subscribers = Subscriber.objects.filter(
                    lists=campaign.subscriber_list,
                    is_active=True
                ).distinct()
                
                if active_subscribers.exists():
                    # Check if this email has already been sent/scheduled to all active subscribers
                    existing_requests = EmailSendRequest.objects.filter(
                        email=email,
                        subscriber__in=active_subscribers,
                        status__in=['sent', 'pending', 'queued']
                    ).values_list('subscriber_id', flat=True)
                    
                    # Find subscribers who haven't received this email yet
                    subscribers_to_schedule = active_subscribers.exclude(id__in=existing_requests)
                    
                    if subscribers_to_schedule.exists():
                        # Calculate scheduled_for time based on email's wait_time and wait_unit
                        # Use 0 as default if wait_time is None, but allow 0 to mean immediate sending
                        wait_time = email.wait_time if email.wait_time is not None else 0
                        wait_unit = email.wait_unit or 'days'
                        
                        send_delay = timedelta(0)
                        if wait_time > 0:
                            if wait_unit == 'seconds':
                                send_delay = timedelta(seconds=wait_time)
                            elif wait_unit == 'minutes':
                                send_delay = timedelta(minutes=wait_time)
                            elif wait_unit == 'hours':
                                send_delay = timedelta(hours=wait_time)
                            elif wait_unit == 'days':
                                send_delay = timedelta(days=wait_time)
                            elif wait_unit == 'weeks':
                                send_delay = timedelta(weeks=wait_time)
                            elif wait_unit == 'months':
                                send_delay = timedelta(days=wait_time * 30)
                        
                        scheduled_for = timezone.now() + send_delay
                        
                        # Create EmailSendRequest for each subscriber who hasn't received this email
                        send_requests = []
                        for subscriber in subscribers_to_schedule:
                            variables = {
                                'first_name': subscriber.first_name or '',
                                'last_name': subscriber.last_name or '',
                                'email': subscriber.email
                            }
                            
                            send_request = EmailSendRequest.objects.create(
                                user=request.user,
                                campaign=campaign,
                                email=email,
                                subscriber=subscriber,
                                subscriber_email=subscriber.email,
                                variables=variables,
                                scheduled_for=scheduled_for,
                                status='pending'
                            )
                            send_requests.append(send_request)
                        
                        logger.info(f"Auto-scheduled updated email '{email.subject}' to {len(send_requests)} active subscribers in campaign '{campaign.name}'")
            
            return Response(serializer.data)
        return Response({'error': 'Validation failed', 'details': serializer.errors}, status=400)
    
    elif request.method == 'DELETE':
        email.delete()
        return Response(status=204)

@login_required
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def test_email(request, campaign_id, email_id):
    """Send a test email."""
    import logging
    campaign = get_object_or_404(Campaign, id=campaign_id, user=request.user)
    email = get_object_or_404(Email, id=email_id, campaign=campaign)
    
    test_email = request.data.get('email')
    if not test_email:
        return Response({'error': _('Email address is required')}, status=400)
    
    variables = request.data.get('variables', {})
    
    # Send test email - try Celery first, fall back to synchronous if unavailable
    import sys
    from django.conf import settings
    from .tasks import send_test_email, _send_test_email_sync
    logger = logging.getLogger(__name__)
    
    # For Windows development, prefer synchronous sending if DEBUG is True
    use_sync = (sys.platform == 'win32' and settings.DEBUG)
    
    if use_sync:
        # On Windows dev, send synchronously by default
        try:
            _send_test_email_sync(str(email.id), test_email, variables)
            return Response({'message': _('Test email sent successfully')})
        except Exception as sync_error:
            return Response({
                'error': _('Failed to send test email: {}').format(str(sync_error))
            }, status=500)
    else:
        # Send test email directly (no Celery)
        try:
            send_test_email(str(email.id), test_email, variables)
            return Response({'message': _('Test email sent successfully')})
        except Exception as e:
            return Response({
                'error': _('Failed to send test email: {}').format(str(e))
            }, status=500)

@api_view(['POST'])
@authentication_classes([BearerTokenAuthentication, TokenAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def send_email(request, campaign_id, email_id):
    """Send an email to a specific subscriber."""
    import logging
    campaign = get_object_or_404(Campaign, id=campaign_id, user=request.user)
    email = get_object_or_404(Email, id=email_id, campaign=campaign)
    
    subscriber_email = request.data.get('email')
    if not subscriber_email:
        return Response({'error': _('Email address is required')}, status=400)
    
    variables = request.data.get('variables', {})
    
    # Check if we should send immediately (bypass wait time)
    send_immediately = request.data.get('send_immediately', False)
    
    # Get or create subscriber and subscribe to campaign's list
    from subscribers.models import Subscriber
    try:
        subscriber, created = Subscriber.objects.get_or_create(
            email=subscriber_email,
            defaults={
                'first_name': variables.get('first_name', ''),
                'last_name': variables.get('last_name', ''),
                'is_active': True
            }
        )
        if not created:
            # Update subscriber info if provided
            updated = False
            if variables.get('first_name') and subscriber.first_name != variables.get('first_name'):
                subscriber.first_name = variables.get('first_name', '')
                updated = True
            if variables.get('last_name') and subscriber.last_name != variables.get('last_name'):
                subscriber.last_name = variables.get('last_name', '')
                updated = True
            if not subscriber.is_active:
                subscriber.is_active = True
                updated = True
            if updated:
                subscriber.save()
        
        # Always subscribe the user to the campaign's list (if campaign has a list)
        if campaign.subscriber_list:
            if not campaign.subscriber_list.subscribers.filter(id=subscriber.id).exists():
                campaign.subscriber_list.subscribers.add(subscriber)
    except Exception as sub_error:
        logger = logging.getLogger(__name__)
        logger.warning(f"Error creating/updating subscriber: {str(sub_error)}")
        # Continue with email sending even if subscriber creation fails
    
    # Create EmailSendRequest record with scheduled time based on email's wait_time
    from campaigns.models import EmailSendRequest
    from django.utils import timezone
    from datetime import timedelta
    from .tasks import _send_single_email_sync
    
    # Calculate scheduled_for time based on email's wait_time and wait_unit
    # If send_immediately is True, schedule for now
    if send_immediately:
        scheduled_for = timezone.now()
    else:
        wait_time = email.wait_time or 1
        wait_unit = email.wait_unit or 'days'
        
        send_delay = timedelta(0)
        if wait_unit == 'minutes':
            send_delay = timedelta(minutes=wait_time)
        elif wait_unit == 'hours':
            send_delay = timedelta(hours=wait_time)
        elif wait_unit == 'days':
            send_delay = timedelta(days=wait_time)
        elif wait_unit == 'weeks':
            send_delay = timedelta(weeks=wait_time)
        elif wait_unit == 'months':
            send_delay = timedelta(days=wait_time * 30)
        
        scheduled_for = timezone.now() + send_delay
    
    send_request = EmailSendRequest.objects.create(
        user=request.user,
        campaign=campaign,
        email=email,
        subscriber=subscriber,
        subscriber_email=subscriber_email,
        variables=variables or {},
        scheduled_for=scheduled_for,
        status='pending'
    )
    
    # Send email synchronously (no Celery)
    logger = logging.getLogger(__name__)
    
    # If scheduled for now or in the past, send immediately
    if scheduled_for <= timezone.now():
        try:
            _send_single_email_sync(str(email.id), subscriber_email, variables, request_id=str(send_request.id))
            send_request.refresh_from_db()
            return Response({
                'message': _('Email sent successfully'),
                'request_id': str(send_request.id),
                'status': send_request.status
            })
        except Exception as sync_error:
            send_request.status = 'failed'
            send_request.error_message = str(sync_error)
            send_request.save(update_fields=['status', 'error_message', 'updated_at'])
            return Response({
                'error': _('Failed to send email: {}').format(str(sync_error))
            }, status=500)
    else:
        # Scheduled for future - cron.py will pick it up
        # Format the scheduled time for the response
        scheduled_display = scheduled_for.strftime('%b %d, %Y %I:%M %p')
        return Response({
            'message': _('Email scheduled for {time}. It will be sent automatically at the scheduled time.').format(time=scheduled_display),
            'request_id': str(send_request.id),
            'scheduled_for': scheduled_for.isoformat(),
            'status': 'pending'
        })

@login_required
def campaign_edit(request, campaign_id):
    """Edit campaign and its templates."""
    from .models import EmailSendRequest
    campaign = get_object_or_404(Campaign, id=campaign_id, user=request.user)
    emails = campaign.emails.all().order_by('order')
    lists = List.objects.filter(user=request.user)
    
    # Get sent emails (status='sent')
    sent_emails = EmailSendRequest.objects.filter(
        campaign=campaign,
        status='sent'
    ).select_related('email', 'subscriber').order_by('-sent_at')
    
    # Get pending emails (status='pending' or 'queued')
    pending_emails = EmailSendRequest.objects.filter(
        campaign=campaign,
        status__in=['pending', 'queued']
    ).select_related('email', 'subscriber').order_by('scheduled_for')
    
    return render(request, 'campaigns/edit.html', {
        'campaign': campaign,
        'emails': emails,
        'lists': lists,
        'sent_emails': sent_emails,
        'pending_emails': pending_emails,
    })

@login_required
def campaign_template(request, campaign_id=None):
    """Render template creation/edit page."""
    campaign = None
    template = None
    show_campaign_modal = False
    
    # Get all user campaigns
    all_campaigns = Campaign.objects.filter(user=request.user).order_by('-created_at')
    
    if campaign_id:
        campaign = get_object_or_404(Campaign, id=campaign_id, user=request.user)
        template_id = request.GET.get('template_id')
        if template_id:
            template = get_object_or_404(Email, id=template_id, campaign=campaign)
    else:
        # If no campaign_id provided from dashboard, show campaign selection modal on save
        if not all_campaigns.exists():
            # No campaigns exist, show error message
            messages.error(request, _('Please create a campaign first before creating email templates.'))
            return redirect('core:dashboard')
        show_campaign_modal = True
    
    # Get user's footers
    from analytics.models import EmailFooter
    user_footers = EmailFooter.objects.filter(user=request.user)
    
    context = {
        'campaign': campaign,
        'template': template,
        'user_footers': user_footers,
        'all_campaigns': all_campaigns,
        'show_campaign_modal': show_campaign_modal,
        'wait_units': [
            ('seconds', _('Seconds')),
            ('minutes', _('Minutes')),
            ('hours', _('Hours')),
            ('days', _('Days')),
            ('weeks', _('Weeks')),
            ('months', _('Months'))
        ],
        'first_name': '{{first_name}}',
        'last_name': '{{last_name}}',
        'full_name': '{{full_name}}',
        'email': '{{email}}'
    }
    return render(request, 'campaigns/template.html', context)

@login_required
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def activate_campaign(request, campaign_id):
    """Activate a campaign."""
    campaign = get_object_or_404(Campaign, id=campaign_id, user=request.user)
    
    if campaign.emails.count() == 0:
        return Response({
            'error': _('Cannot activate a campaign with no emails')
        }, status=400)
    
    campaign.is_active = True
    campaign.save()
    
    return Response({
        'message': _('Campaign activated successfully')
    })

@login_required
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def deactivate_campaign(request, campaign_id):
    """Deactivate a campaign."""
    campaign = get_object_or_404(Campaign, id=campaign_id, user=request.user)
    campaign.is_active = False
    campaign.save()
    
    return Response({
        'message': _('Campaign deactivated successfully')
    })

@login_required
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reorder_emails(request, campaign_id):
    """Reorder emails in a campaign."""
    campaign = get_object_or_404(Campaign, id=campaign_id, user=request.user)
    
    email_id = request.data.get('email_id')
    direction = request.data.get('direction')  # 'up' or 'down'
    
    if not email_id or not direction:
        return Response({'error': _('Email ID and direction are required')}, status=400)
    
    try:
        email = Email.objects.get(id=email_id, campaign=campaign)
    except Email.DoesNotExist:
        return Response({'error': _('Email not found')}, status=404)
    
    emails = list(campaign.emails.all().order_by('order'))
    current_index = None
    
    # Find current email index
    for i, e in enumerate(emails):
        if e.id == email.id:
            current_index = i
            break
    
    if current_index is None:
        return Response({'error': _('Email not found in campaign')}, status=404)
    
    # Determine new index
    if direction == 'up' and current_index > 0:
        new_index = current_index - 1
    elif direction == 'down' and current_index < len(emails) - 1:
        new_index = current_index + 1
    else:
        return Response({'error': _('Cannot move email in that direction')}, status=400)
    
    # Swap orders
    emails[current_index], emails[new_index] = emails[new_index], emails[current_index]
    
    # Update orders
    for i, e in enumerate(emails):
        e.order = i
        e.save(update_fields=['order'])
    
    return Response({'message': _('Email order updated successfully')})

@login_required
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_contacts(request):
    """Upload contacts from a CSV or Excel file."""
    if 'file' not in request.FILES:
        return Response({'error': _('No file uploaded')}, status=400)
    
    file = request.FILES['file']
    if not file.name.endswith(('.csv', '.xls', '.xlsx')):
        return Response({'error': _('Invalid file format. Please upload a CSV or Excel file.')}, status=400)
    
    try:
        if file.name.endswith('.csv'):
            # Read CSV file
            file.seek(0)  # Reset file pointer
            decoded_file = file.read().decode('utf-8-sig')
            reader = csv.DictReader(decoded_file.splitlines())
            rows = list(reader)
            columns = reader.fieldnames or []
        else:
            # Read Excel file
            workbook = load_workbook(filename=file, read_only=True, data_only=True)
            worksheet = workbook.active
            rows = []
            columns = None
            for idx, row in enumerate(worksheet.iter_rows(values_only=True)):
                if idx == 0:
                    columns = [str(cell) if cell else f'Column_{i+1}' for i, cell in enumerate(row)]
                else:
                    row_dict = {}
                    for i, cell in enumerate(row):
                        if i < len(columns):
                            row_dict[columns[i]] = cell
                    rows.append(row_dict)
            workbook.close()
        
        # Validate required columns
        required_columns = ['email']
        missing_columns = [col for col in required_columns if col not in columns]
        if missing_columns:
            return Response({
                'error': _('Missing required columns: {}').format(', '.join(missing_columns))
            }, status=400)
        
        # Process the data
        contacts = []
        for row in rows:
            contact = {
                'email': row.get('email', ''),
                'first_name': row.get('first_name', '') or '',
                'last_name': row.get('last_name', '') or '',
                'full_name': row.get('full_name', '') or ''
            }
            contacts.append(contact)
        
        return Response({
            'message': _('File processed successfully'),
            'contacts': contacts
        })
        
    except Exception as e:
        return Response({
            'error': _('Error processing file: {}').format(str(e))
        }, status=400)

@login_required
def campaign_create(request):
    """Render campaign creation page."""
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        subscriber_list_id = request.POST.get('subscriber_list')
        
        if name:
            subscriber_list = None
            if subscriber_list_id:
                subscriber_list = get_object_or_404(List, id=subscriber_list_id, user=request.user)
            
            campaign = Campaign.objects.create(
                user=request.user,
                name=name,
                description=description or '',
                subscriber_list=subscriber_list
            )
            
            messages.success(request, _('Campaign created successfully!'))
            return redirect('campaigns:edit', campaign_id=campaign.id)
        else:
            messages.error(request, _('Please provide a campaign name.'))
    
    return render(request, 'campaigns/create.html', {
        'subscriber_lists': List.objects.filter(user=request.user)
    })


@login_required
def template_revision_view(request):
    """Simple web view for template revision and editing.

    - GET: select a campaign and email to preview
    - POST: save an edited draft of an email's HTML body
    """
    campaigns = Campaign.objects.filter(user=request.user)
    campaign = None
    emails = []
    email = None

    # Allow selecting campaign/email via querystring or form
    campaign_id = request.GET.get('campaign_id') or request.POST.get('campaign_id')
    if campaign_id:
        try:
            campaign = Campaign.objects.get(id=campaign_id, user=request.user)
            emails = campaign.emails.all().order_by('order')

            email_id = request.GET.get('email_id') or request.POST.get('email_id')
            if email_id:
                try:
                    email = campaign.emails.get(id=email_id)
                except Email.DoesNotExist:
                    email = None
        except (Campaign.DoesNotExist, ValueError):
            # Invalid campaign_id, just show empty state
            campaign = None
            emails = []

    if request.method == 'POST' and email:
        draft = request.POST.get('draft', '')
        email.body_html = draft
        email.save(update_fields=['body_html', 'updated_at'])
        messages.success(request, 'Draft saved to database.')
        # Redirect back to the same selection
        url = reverse('campaigns:template_revision')
        params = f'?campaign_id={campaign.id}&email_id={email.id}'
        return redirect(url + params)

    context = {
        'campaigns': campaigns,
        'campaign': campaign,
        'emails': emails,
        'email': email,
    }
    try:
        return render(request, 'campaigns/template_revision.html', context)
    except TemplateDoesNotExist:
        # Support templates placed directly under a templates/ folder without app subdir
        return render(request, 'template_revision.html', context)


@login_required
def drafts_view(request):
    """View saved email drafts with campaign filtering."""
    campaigns = Campaign.objects.filter(user=request.user).order_by('-created_at')
    campaign = None
    drafts = []
    
    # Get user's footers for the send modal
    from analytics.models import EmailFooter
    user_footers = EmailFooter.objects.filter(user=request.user)
    
    # Allow selecting campaign via querystring
    campaign_id = request.GET.get('campaign_id')
    if campaign_id:
        try:
            campaign = Campaign.objects.get(id=campaign_id, user=request.user)
            # Get all emails (drafts) for this campaign
            drafts = campaign.emails.all().order_by('order')
        except (Campaign.DoesNotExist, ValueError):
            campaign = None
            drafts = []
    
    context = {
        'campaigns': campaigns,
        'campaign': campaign,
        'drafts': drafts,
        'user_footers': user_footers,
    }
    return render(request, 'campaigns/drafts.html', context)


@login_required
def campaign_analysis_view(request):
    """Render campaign analysis similar to the Streamlit view.

    Shows basic metrics, monthly aggregates and per-email breakdown.
    """
    campaigns = Campaign.objects.filter(user=request.user)
    campaign = None
    by_month = {}
    email_rows = []

    campaign_id = request.GET.get('campaign_id')
    if campaign_id:
        campaign = get_object_or_404(Campaign, id=campaign_id, user=request.user)

        events = EmailEvent.objects.filter(email__campaign=campaign)
        from collections import defaultdict
        by_month = defaultdict(lambda: {'opened': 0, 'clicked': 0, 'sent': 0, 'bounces': 0, 'unsubscribes': 0, 'complaints': 0})
        for ev in events:
            month = ev.created_at.strftime('%Y-%m')
            if ev.event_type == 'opened':
                by_month[month]['opened'] += 1
            elif ev.event_type == 'clicked':
                by_month[month]['clicked'] += 1
            elif ev.event_type == 'sent':
                by_month[month]['sent'] += 1
            elif ev.event_type == 'bounced':
                by_month[month]['bounces'] += 1
            elif ev.event_type == 'unsubscribed':
                by_month[month]['unsubscribes'] += 1
            elif ev.event_type == 'complained':
                by_month[month]['complaints'] += 1

        # Per-email breakdown with unique opens
        for e in campaign.emails.all():
            sent = EmailEvent.objects.filter(email=e, event_type='sent').count()
            # Count unique opens (distinct subscriber emails that opened)
            unique_opens = EmailEvent.objects.filter(
                email=e, 
                event_type='opened'
            ).values('subscriber_email').distinct().count()
            # Total clicks
            clicked = EmailEvent.objects.filter(email=e, event_type='clicked').count()
            open_rate = (unique_opens / sent * 100) if sent > 0 else 0
            click_rate = (clicked / sent * 100) if sent > 0 else 0
            email_rows.append({
                'subject': e.subject,
                'sent': sent,
                'opened': unique_opens,  # Use unique opens for display
                'clicked': clicked,
                'open_rate': round(open_rate, 2),
                'click_rate': round(click_rate, 2)
            })

    # Convert defaultdict to a sorted list of (month, values) tuples for the template
    # This avoids relying on the template dictsort filter which may behave differently
    # across Django versions or custom template configurations.
    month_items = []
    if by_month:
        # sort by month string (YYYY-MM) ascending
        month_items = sorted(by_month.items())

    return render(request, 'campaigns/campaign_analysis.html', {
        'campaigns': campaigns,
        'campaign': campaign,
        'by_month_items': month_items,
        'email_rows': email_rows,
    })


@login_required
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def campaign_stats_api(request, campaign_id):
    """API endpoint for real-time campaign statistics."""
    campaign = get_object_or_404(Campaign, id=campaign_id, user=request.user)
    
    logger.info(f"Fetching stats for campaign {campaign_id} - {campaign.name}")
    
    # Get real-time counts directly from EmailEvent for accuracy
    sent_count = EmailEvent.objects.filter(
        email__campaign=campaign, 
        event_type='sent'
    ).count()
    logger.info(f"Campaign {campaign_id}: Found {sent_count} sent events")
    bounce_count = EmailEvent.objects.filter(
        email__campaign=campaign, 
        event_type='bounced'
    ).count()
    unsubscribe_count = EmailEvent.objects.filter(
        email__campaign=campaign, 
        event_type='unsubscribed'
    ).count()
    complaint_count = EmailEvent.objects.filter(
        email__campaign=campaign, 
        event_type='complained'
    ).count()
    
    # Get opens and clicks from campaign model (these are incremented via signals)
    campaign.refresh_from_db()
    open_count = campaign.open_count
    click_count = campaign.click_count
    
    # Calculate rates
    delivered_count = sent_count - bounce_count
    delivery_rate = (delivered_count / sent_count * 100) if sent_count > 0 else 0
    open_rate = (open_count / sent_count * 100) if sent_count > 0 else 0
    click_rate = (click_count / sent_count * 100) if sent_count > 0 else 0
    
    # Get overall campaign stats
    overall_stats = {
        'sent_count': sent_count,
        'open_count': open_count,
        'click_count': click_count,
        'open_rate': round(open_rate, 2),
        'click_rate': round(click_rate, 2),
        'delivery_rate': round(delivery_rate, 2),
        'bounce_count': bounce_count,
        'unsubscribe_count': unsubscribe_count,
        'complaint_count': complaint_count,
    }
    
    # Get per-email stats
    email_stats = []
    for email in campaign.emails.all().order_by('order'):
        sent = EmailEvent.objects.filter(email=email, event_type='sent').count()
        opened = EmailEvent.objects.filter(email=email, event_type='opened').count()
        clicked = EmailEvent.objects.filter(email=email, event_type='clicked').count()
        unique_opens = EmailEvent.objects.filter(
            email=email, 
            event_type='opened'
        ).values('subscriber_email').distinct().count()
        
        email_stats.append({
            'id': str(email.id),
            'subject': email.subject,
            'order': email.order,
            'sent': sent,
            'opened': opened,
            'clicked': clicked,
            'unique_opens': unique_opens,
            'open_rate': round((opened / sent * 100) if sent > 0 else 0, 2),
            'click_rate': round((clicked / sent * 100) if sent > 0 else 0, 2),
        })
    
    # Get recent events (last 50)
    recent_events = EmailEvent.objects.filter(
        email__campaign=campaign
    ).select_related('email').order_by('-created_at')[:50]
    
    events_list = [{
        'id': str(event.id),
        'email_subject': event.email.subject,
        'subscriber_email': event.subscriber_email,
        'event_type': event.event_type,
        'link_clicked': event.link_clicked,
        'created_at': event.created_at.isoformat(),
    } for event in recent_events]
    
    return Response({
        'campaign': {
            'id': str(campaign.id),
            'name': campaign.name,
            'is_active': campaign.is_active,
        },
        'overall_stats': overall_stats,
        'email_stats': email_stats,
        'recent_events': events_list,
    })


@login_required
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_email_with_ai(request, campaign_id):
    """Generate email subject and body using Hugging Face API."""
    campaign = get_object_or_404(Campaign, id=campaign_id, user=request.user)
    
    subject_topic = request.data.get('subject_topic')
    recipient = request.data.get('recipient', 'subscriber')
    tone = request.data.get('tone', 'professional')
    length = request.data.get('length', 'medium')
    context = request.data.get('context', '')
    email_id = request.data.get('email_id')
    
    if not subject_topic:
        return Response({'error': _('Subject topic is required')}, status=400)
    
    try:
        # Generate content using AI
        generated = generate_email_content(
            subject=subject_topic,
            recipient=recipient,
            tone=tone,
            length=length,
            context=context
        )
        
        # If an email_id is provided, save to database
        if email_id:
            try:
                email = Email.objects.get(id=email_id, campaign=campaign)
                email.subject = generated.get('subject', email.subject)
                email.body_html = generated.get('body_html', email.body_html)
                email.save(update_fields=['subject', 'body_html', 'updated_at'])
                
                # Store AI analysis record
                ai_analysis, created = EmailAIAnalysis.objects.get_or_create(email=email)
                ai_analysis.generated_subject = generated.get('subject')
                ai_analysis.generated_body_html = generated.get('body_html')
                ai_analysis.generation_prompt = f"Topic: {subject_topic}, Tone: {tone}, Length: {length}"
                ai_analysis.save()
                
                return Response({
                    'message': _('Email generated and saved successfully'),
                    'subject': generated.get('subject'),
                    'body_html': generated.get('body_html'),
                    'saved': True
                })
            except Email.DoesNotExist:
                return Response({'error': _('Email not found')}, status=404)
        else:
            # Just return the generated content
            return Response({
                'message': _('Email generated successfully'),
                'subject': generated.get('subject'),
                'body_html': generated.get('body_html'),
                'saved': False
            })
    
    except ValueError as e:
        logger.error(f"AI email generation error (ValueError): {str(e)}")
        return Response({
            'error': str(e),
            'hint': 'Please ensure Ollama is running (ollama serve) and the llama3.1:8b model is available (ollama pull llama3.1:8b)'
        }, status=400)
    except Exception as e:
        import traceback
        logger.error(f"AI email generation error (Exception): {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        return Response({
            'error': _('Failed to generate email: {}').format(str(e)),
            'detail': str(e)
        }, status=500)


@login_required
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def revise_email_with_ai(request):
    """Revise email content for grammar and clarity using Hugging Face API."""
    from .ai_utils import revise_email_content
    
    email_text = request.data.get('email_text', '')
    
    if not email_text or not email_text.strip():
        return Response({'error': _('Email text is required')}, status=400)
    
    try:
        result = revise_email_content(email_text)
        
        if result.get('success'):
            return Response({
                'message': _('Email revised successfully'),
                'revised_text': result.get('revised_text'),
                'success': True
            })
        else:
            return Response({
                'error': result.get('error', _('Revision failed')),
                'success': False
            }, status=500)
    
    except Exception as e:
        return Response({
            'error': _('Failed to revise email: {}').format(str(e)),
            'success': False
        }, status=500)


@login_required
@csrf_exempt
def search_templates(request):
    """Search through all email templates and campaigns."""
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    if not request.user.is_authenticated:
        return JsonResponse({'results': []}, status=401)
    
    query = request.GET.get('q', '').strip()
    
    if not query or len(query) < 2:
        return JsonResponse({'results': []})
    
    from django.db.models import Q
    import re
    
    results = []
    
    # Search campaigns
    campaigns = Campaign.objects.filter(
        user=request.user
    ).filter(
        Q(name__icontains=query) | Q(description__icontains=query)
    )[:5]
    
    for campaign in campaigns:
        results.append({
            'type': 'campaign',
            'id': str(campaign.id),
            'name': campaign.name,
            'description': campaign.description or '',
            'is_active': campaign.is_active,
            'emails_count': campaign.emails_count,
            'url': reverse('campaigns:edit', args=[campaign.id])
        })
    
    # Search email templates
    emails = Email.objects.filter(
        campaign__user=request.user
    ).filter(
        Q(subject__icontains=query) | Q(body_html__icontains=query) | Q(body_text__icontains=query)
    ).select_related('campaign').distinct()[:10]
    
    for email in emails:
        # Create preview by stripping HTML tags
        preview = re.sub(r'<[^>]+>', '', email.body_html or email.body_text or '')
        preview = preview.strip()[:100] + ('...' if len(preview) > 100 else '')
        
        template_url = reverse('campaigns:template', args=[email.campaign.id])
        template_url += f'?template_id={email.id}'
        
        results.append({
            'type': 'template',
            'id': str(email.id),
            'subject': email.subject or 'Untitled Email',
            'preview': preview,
            'campaign_id': str(email.campaign.id),
            'campaign_name': email.campaign.name,
            'order': email.order,
            'url': template_url
        })
    
    return JsonResponse({'results': results})


@login_required
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def send_pending_email(request, request_id):
    """Send a specific pending email immediately."""
    from .models import EmailSendRequest
    from .tasks import _send_single_email_sync
    from django.utils import timezone
    
    # Get the email send request
    send_request = get_object_or_404(EmailSendRequest, id=request_id, user=request.user)
    
    # Check if already sent
    if send_request.status == 'sent':
        return Response({'error': 'Email already sent'}, status=400)
    
    # Check if failed
    if send_request.status == 'failed':
        return Response({'error': 'Email previously failed. Cannot resend.'}, status=400)
    
    try:
        # Update status to queued
        send_request.status = 'queued'
        send_request.save(update_fields=['status', 'updated_at'])
        
        # Send the email synchronously
        _send_single_email_sync(
            email_id=send_request.email.id,
            subscriber_email=send_request.subscriber_email,
            variables=send_request.variables,
            request_id=send_request.id
        )
        
        # Refresh from DB to get updated status
        send_request.refresh_from_db()
        
        # If status is still not sent, there might be an error
        if send_request.status != 'sent':
            return Response({
                'error': f'Email sending failed: {send_request.error_message}'
            }, status=500)
        
        # Update campaign statistics
        campaign = send_request.campaign
        campaign.sent_count += 1
        campaign.save(update_fields=['sent_count', 'updated_at'])
        
        return Response({
            'success': True,
            'message': 'Email sent successfully',
            'sent_at': send_request.sent_at.isoformat() if send_request.sent_at else None
        })
    
    except Exception as e:
        logger.error(f"Error sending pending email {request_id}: {str(e)}")
        send_request.status = 'failed'
        send_request.error_message = str(e)
        send_request.save(update_fields=['status', 'error_message', 'updated_at'])
        return Response({'error': str(e)}, status=500)

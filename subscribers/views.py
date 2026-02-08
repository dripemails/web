from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils.translation import gettext as _
from django.core.paginator import Paginator
from django.db.models import Q, Value, CharField
from django.db.models.functions import Coalesce, Concat, Trim
from django.conf import settings
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .models import List, Subscriber, CustomField, CustomValue
from .serializers import ListSerializer, SubscriberSerializer
from campaigns.models import Campaign
import csv
import json
import math
import time
from openpyxl import load_workbook
from datetime import datetime
from django.db import connection
from django.db.utils import OperationalError


def _json_safe(val):
    """Convert NaN/Inf and non-JSON types to JSON-serializable values."""
    if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
        return None
    if isinstance(val, dict):
        return {k: _json_safe(v) for k, v in val.items()}
    if isinstance(val, (list, tuple)):
        return [_json_safe(v) for v in val]
    if hasattr(val, 'item'):  # numpy scalar
        try:
            return val.item()
        except (ValueError, AttributeError):
            return None
    return val


# Column name aliases for import: file header (normalized) -> our internal key.
IMPORT_COLUMN_ALIASES = {
    'email': ['email', 'email address', 'e-mail', 'emailaddress', 'mail'],
    'first_name': ['first name', 'firstname', 'first', 'given name', 'givenname'],
    'last_name': ['last name', 'lastname', 'last', 'surname', 'family name', 'familyname'],
    'company': ['company', 'organization', 'org'],
}


def _normalize_header(name):
    if name is None:
        return ''
    return ' '.join(str(name).strip().lower().split())


def _build_import_column_map(columns):
    """Map file column names to our keys (email, first_name, last_name). Returns dict: our_key -> file_column_name."""
    alias_to_key = {}
    for key, aliases in IMPORT_COLUMN_ALIASES.items():
        for a in aliases:
            alias_to_key[a] = key
    column_map = {}
    for file_col in (columns or []):
        norm = _normalize_header(file_col)
        if norm in alias_to_key:
            key = alias_to_key[norm]
            if key not in column_map:
                column_map[key] = file_col
    return column_map


def _cell_str(val, default=''):
    """Get a string from a cell value; handle pandas NaN and None."""
    if val is None:
        return default
    if hasattr(val, 'item') and hasattr(val, 'dtype'):
        try:
            val = val.item()
        except (ValueError, AttributeError):
            return default
    if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
        return default
    s = str(val).strip()
    return s if s and s.lower() != 'nan' else default

@login_required
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def list_list_create(request):
    """List all subscriber lists or create a new one."""
    if request.method == 'GET':
        lists = List.objects.filter(user=request.user)
        serializer = ListSerializer(lists, many=True)
        return Response(serializer.data)
    
    serializer = ListSerializer(data=request.data, context={'request': request})
    if serializer.is_valid():
        list_obj = serializer.save(user=request.user)
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)

@login_required
@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def list_detail(request, pk):
    """Retrieve, update or delete a subscriber list."""
    list_obj = get_object_or_404(List, id=pk, user=request.user)
    
    if request.method == 'GET':
        serializer = ListSerializer(list_obj)
        return Response(serializer.data)
    
    elif request.method == 'PUT':
        serializer = ListSerializer(list_obj, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)
    
    elif request.method == 'DELETE':
        list_obj.delete()
        return Response(status=204)

@login_required
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def subscriber_list_create(request):
    """List all subscribers or create a new one."""
    if request.method == 'GET':
        list_id = request.query_params.get('list_id')
        if list_id:
            list_obj = get_object_or_404(List, id=list_id, user=request.user)
            subscribers = list_obj.subscribers.all()
        else:
            subscribers = Subscriber.objects.filter(lists__user=request.user).distinct()
        
        serializer = SubscriberSerializer(subscribers, many=True, context={'request': request})
        return Response(serializer.data)
    
    list_id = request.data.get('list_id') or None
    list_obj = None

    if list_id:
        try:
            list_obj = List.objects.get(id=list_id, user=request.user)
        except List.DoesNotExist:
            return Response({'list_id': _('Selected list does not exist.')}, status=404)
 
    serializer = SubscriberSerializer(
        data=request.data,
        context={'request': request, 'list_obj': list_obj, 'list_id': list_id}
    )
    if serializer.is_valid():
        try:
            subscriber = serializer.save()
            
            # If no list was provided, ensure subscriber is added to at least one list
            # so they appear on the subscribers page
            if not list_obj and subscriber.lists.count() == 0:
                # Create or get a default list for the user
                default_list, created = List.objects.get_or_create(
                    user=request.user,
                    name='Default List',
                    defaults={
                        'description': 'Default list for subscribers without a specific list assignment'
                    }
                )
                subscriber.lists.add(default_list)
            
            # Update the serializer instance with the saved object to get the response data
            serializer.instance = subscriber
            return Response(serializer.data, status=201)
        except Exception as e:
            import traceback
            import logging
            error_trace = traceback.format_exc()
            logger = logging.getLogger(__name__)
            logger.error(f"Error creating subscriber: {str(e)}\n{error_trace}")
            return Response({
                'error': str(e),
                'traceback': error_trace if settings.DEBUG else None
            }, status=500)
    return Response(serializer.errors, status=400)

@login_required
@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def subscriber_detail(request, pk):
    """Retrieve, update or delete a subscriber."""
    subscriber = get_object_or_404(Subscriber, id=pk, lists__user=request.user)
    
    if request.method == 'GET':
        serializer = SubscriberSerializer(subscriber, context={'request': request})
        return Response(serializer.data)
    
    elif request.method == 'PUT':
        serializer = SubscriberSerializer(
            subscriber, 
            data=request.data,
            context={'request': request}
        )
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)
    
    elif request.method == 'DELETE':
        subscriber.delete()
        return Response(status=204)

@login_required
def import_subscribers(request):
    """Render subscriber import page."""
    campaigns = Campaign.objects.filter(user=request.user)
    lists = List.objects.filter(user=request.user)
    context = {
        'campaigns': campaigns,
        'lists': lists
    }
    return render(request, 'subscribers/import.html', context)

@login_required
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def process_import(request):
    """Process the uploaded file and import subscribers."""
    file = request.FILES.get('file')
    list_id = request.data.get('list_id')
    campaign_id = request.data.get('campaign_id')
    mappings = json.loads(request.data.get('mappings', '{}'))
    
    if not file or not list_id:
        return Response({
            'error': _('Please provide both a file and select a list')
        }, status=400)
    
    try:
        # Read file based on extension
        if file.name.endswith('.csv'):
            # Read CSV file
            file.seek(0)  # Reset file pointer
            decoded_file = file.read().decode('utf-8-sig')  # Handle BOM
            reader = csv.DictReader(decoded_file.splitlines())
            rows = list(reader)
            columns = reader.fieldnames or []
        elif file.name.endswith(('.xls', '.xlsx')):
            # Read Excel file - use pandas for better .xls/.xlsx support
            try:
                import pandas as pd
                # Read Excel file into DataFrame
                df = pd.read_excel(file, engine='openpyxl' if file.name.endswith('.xlsx') else 'xlrd')
                # Convert to list of dictionaries
                rows = df.to_dict('records')
                # Get column names
                columns = list(df.columns)
            except ImportError:
                # Fallback to openpyxl if pandas not available (only works for .xlsx)
                if file.name.endswith('.xls'):
                    return Response({
                        'error': _('Please install pandas and xlrd to support .xls files, or convert to .xlsx format')
                    }, status=400)
                workbook = load_workbook(filename=file, read_only=True, data_only=True)
                worksheet = workbook.active
                rows = []
                columns = None
                for idx, row in enumerate(worksheet.iter_rows(values_only=True)):
                    if idx == 0:
                        # First row is headers
                        columns = [str(cell) if cell else f'Column_{i+1}' for i, cell in enumerate(row)]
                    else:
                        # Convert row to dict
                        row_dict = {}
                        for i, cell in enumerate(row):
                            if i < len(columns):
                                row_dict[columns[i]] = cell
                        rows.append(row_dict)
                workbook.close()
        else:
            return Response({
                'error': _('Unsupported file format')
            }, status=400)
        
        # Auto-map columns when mappings are empty or don't match (e.g. "Email Address" -> email)
        column_map = _build_import_column_map(columns)
        email_column = (mappings.get('email') or column_map.get('email') or 'email').strip() or column_map.get('email')
        first_name_col = (mappings.get('first_name') or column_map.get('first_name') or '').strip() or column_map.get('first_name')
        last_name_col = (mappings.get('last_name') or column_map.get('last_name') or '').strip() or column_map.get('last_name')
        
        file_columns = columns or (list(rows[0].keys()) if rows else [])
        if not email_column or email_column not in file_columns:
            return Response({
                'error': _('Could not find an email column. Your file must have a column named "Email", "Email Address", or similar.')
            }, status=400)
        
        skip_cols = {c for c in (email_column, first_name_col, last_name_col) if c}
        
        # Retry on SQLite "database is locked" (common when many rows are imported)
        max_retries = 8
        for attempt in range(max_retries):
            try:
                connection.close()  # Release any stale connection before retry
                subscriber_list = List.objects.get(id=list_id, user=request.user)
                campaign = None
                if campaign_id:
                    campaign = Campaign.objects.get(id=campaign_id, user=request.user)
                
                imported_count = 0
                skipped_count = 0
                
                for row in rows:
                    email = _cell_str(row.get(email_column))
                    if not email:
                        skipped_count += 1
                        continue
                    
                    first_name = _cell_str(row.get(first_name_col)) if first_name_col else ''
                    last_name = _cell_str(row.get(last_name_col)) if last_name_col else ''
                    
                    subscriber, created = Subscriber.objects.get_or_create(
                        email=email,
                        defaults={
                            'first_name': first_name,
                            'last_name': last_name,
                            'is_active': True
                        }
                    )
                    
                    if subscriber_list not in subscriber.lists.all():
                        subscriber.lists.add(subscriber_list)
                    
                    for column in (columns or list(row.keys())):
                        if column in skip_cols:
                            continue
                        value = _cell_str(row.get(column))
                        if not value:
                            continue
                        field, field_created = CustomField.objects.get_or_create(
                            user=request.user,
                            key=column.lower().replace(' ', '_'),
                            defaults={'name': column}
                        )
                        CustomValue.objects.update_or_create(
                            subscriber=subscriber,
                            field=field,
                            defaults={'value': value}
                        )
                    
                    imported_count += 1
                
                return Response({
                    'message': _('Successfully imported %(imported)d subscribers (%(skipped)d skipped)') % {
                        'imported': imported_count,
                        'skipped': skipped_count
                    }
                })
            except OperationalError as e:
                err_msg = str(e).lower()
                if 'locked' in err_msg or 'database is locked' in err_msg:
                    if attempt < max_retries - 1:
                        time.sleep(0.15 * (attempt + 1))
                        continue
                    return Response({
                        'error': _('The database is busy. Please try again in a moment (e.g. wait a few seconds and re-upload).')
                    }, status=503)
                raise
        
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=400)

@login_required
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def validate_file(request):
    """Validate the uploaded file and return column names."""
    file = request.FILES.get('file')
    if not file:
        return Response({'error': _('No file provided')}, status=400)
    
    try:
        # Read first few rows to get columns
        if file.name.endswith('.csv'):
            file.seek(0)  # Reset file pointer
            decoded_file = file.read().decode('utf-8-sig')
            reader = csv.DictReader(decoded_file.splitlines())
            columns = list(reader.fieldnames or [])
            # Get first 5 rows for preview
            preview = []
            for i, row in enumerate(reader):
                if i >= 5:
                    break
                preview.append(row)
        elif file.name.endswith(('.xls', '.xlsx')):
            # Read Excel file - use pandas for better .xls/.xlsx support
            try:
                import pandas as pd
                # Read Excel file into DataFrame
                df = pd.read_excel(file, engine='openpyxl' if file.name.endswith('.xlsx') else 'xlrd', nrows=6)
                # Get column names
                columns = list(df.columns)
                # Get first 5 rows for preview
                preview = df.head(5).to_dict('records')
            except ImportError:
                # Fallback to openpyxl if pandas not available (only works for .xlsx)
                if file.name.endswith('.xls'):
                    return Response({
                        'error': _('Please install pandas and xlrd to support .xls files, or convert to .xlsx format')
                    }, status=400)
                workbook = load_workbook(filename=file, read_only=True, data_only=True)
                worksheet = workbook.active
                columns = []
                preview = []
                
                for idx, row in enumerate(worksheet.iter_rows(values_only=True)):
                    if idx == 0:
                        columns = [str(cell) if cell else f'Column_{i+1}' for i, cell in enumerate(row)]
                    elif idx <= 5:
                        row_dict = {}
                        for i, cell in enumerate(row):
                            if i < len(columns):
                                row_dict[columns[i]] = cell
                        preview.append(row_dict)
                    else:
                        break
                workbook.close()
        else:
            return Response({'error': _('Unsupported file format')}, status=400)
        
        # Sanitize so NaN/Inf from pandas don't break JSON serialization
        preview = _json_safe(preview)
        columns = [str(c) if c not in (None, '') else f'Column_{i+1}' for i, c in enumerate(_json_safe(columns))]
        return Response({
            'columns': columns,
            'preview': preview
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=400)


@login_required
def subscriber_directory(request):
    """Display the subscribers for the authenticated user with search and pagination."""
    query = request.GET.get('q', '').strip()

    subscribers = (
        Subscriber.objects.filter(lists__user=request.user)
        .distinct()
        .prefetch_related('lists', 'lists__campaigns')
        .annotate(
            annotated_full_name=Trim(
                Concat(
                    Coalesce('first_name', Value('')),
                    Value(' '),
                    Coalesce('last_name', Value('')),
                    output_field=CharField(),
                )
            )
        )
        .order_by('-created_at')
    )

    if query:
        subscribers = subscribers.filter(
            Q(first_name__icontains=query)
            | Q(last_name__icontains=query)
            | Q(email__icontains=query)
            | Q(annotated_full_name__icontains=query)
        )

    paginator = Paginator(subscribers, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page_obj,
        'query': query,
        'total_count': paginator.count,
    }
    return render(request, 'subscribers/list.html', context)


@login_required
def list_view_edit(request, pk):
    """View and edit a subscriber list with its subscribers."""
    list_obj = get_object_or_404(List, id=pk, user=request.user)
    
    # Handle POST for editing list
    if request.method == 'POST':
        # Check if this is a list edit or subscriber add
        if 'subscriber_action' in request.POST and request.POST.get('subscriber_action') == 'add':
            # Handle adding subscriber to list
            email = request.POST.get('email', '').strip()
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            
            if not email:
                messages.error(request, _('Email is required'))
            else:
                subscriber, created = Subscriber.objects.get_or_create(
                    email=email,
                    defaults={
                        'first_name': first_name,
                        'last_name': last_name,
                        'is_active': True
                    }
                )
                
                if not created:
                    # Update existing subscriber if name fields provided
                    updated = False
                    if first_name and subscriber.first_name != first_name:
                        subscriber.first_name = first_name
                        updated = True
                    if last_name and subscriber.last_name != last_name:
                        subscriber.last_name = last_name
                        updated = True
                    if not subscriber.is_active:
                        subscriber.is_active = True
                        updated = True
                    if updated:
                        subscriber.save()
                
                # Add to list if not already in it
                if list_obj not in subscriber.lists.all():
                    subscriber.lists.add(list_obj)
                    messages.success(request, _('Subscriber added to list successfully!'))
                else:
                    messages.info(request, _('Subscriber is already in this list'))
                
                return redirect('subscribers:list-view-edit', pk=pk)
        else:
            # Handle list edit
            list_obj.name = request.POST.get('name', list_obj.name)
            list_obj.description = request.POST.get('description', list_obj.description)
            list_obj.save()
            messages.success(request, _('List updated successfully!'))
            return redirect('subscribers:list-view-edit', pk=pk)
    
    # Get subscribers in this list
    subscribers = list_obj.subscribers.all().prefetch_related('lists').order_by('-created_at')
    
    # Get campaigns using this list
    campaigns = Campaign.objects.filter(subscriber_list=list_obj, user=request.user)
    
    # Pagination for subscribers
    paginator = Paginator(subscribers, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    
    context = {
        'list_obj': list_obj,
        'page_obj': page_obj,
        'campaigns': campaigns,
        'total_count': paginator.count,
    }
    return render(request, 'subscribers/list_view_edit.html', context)


@login_required
def add_subscriber(request):
    """Render the add subscriber page with list and campaign data."""
    from django.urls import reverse
    
    lists = List.objects.filter(user=request.user).order_by('name')
    campaigns = Campaign.objects.filter(user=request.user).prefetch_related('emails').order_by('name')

    campaign_data = [
        {
            'id': str(campaign.id),
            'name': campaign.name,
            'emails': [
                {
                    'id': str(email.id),
                    'subject': email.subject,
                }
                for email in campaign.emails.all()
            ],
        }
        for campaign in campaigns
    ]

    # Build absolute URLs for API endpoints to avoid language prefix issues
    scheme = 'https' if request.is_secure() else 'http'
    host = request.get_host()
    api_subscriber_create_path = reverse('api-subscriber-create')
    api_send_email_path = reverse('api-send-email')
    api_subscriber_create_url = f"{scheme}://{host}{api_subscriber_create_path}"
    api_send_email_url = f"{scheme}://{host}{api_send_email_path}"

    context = {
        'lists': lists,
        'campaigns': campaigns,
        'campaign_data_json': json.dumps(campaign_data),
        'api_subscriber_create_url': api_subscriber_create_url,
        'api_send_email_url': api_send_email_url,
    }
    return render(request, 'subscribers/add.html', context)


@login_required
def export_subscribers_csv(request):
    """Export all subscribers to CSV file."""
    # Get all subscribers for the user
    subscribers = (
        Subscriber.objects.filter(lists__user=request.user)
        .distinct()
        .prefetch_related('lists')
        .order_by('-created_at')
    )
    
    # Create the HttpResponse object with CSV header
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="subscribers_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    # Create CSV writer
    writer = csv.writer(response)
    
    # Write header row
    writer.writerow([
        'Email',
        'First Name',
        'Last Name',
        'Full Name',
        'Lists',
        'Campaigns',
        'Status',
        'Confirmed',
        'Joined Date'
    ])
    
    # Write subscriber data
    for subscriber in subscribers:
        # Get lists as comma-separated string
        lists = ', '.join([list_obj.name for list_obj in subscriber.lists.all()])
        
        # Get campaigns from all lists
        campaign_names = []
        for list_obj in subscriber.lists.all():
            for campaign in list_obj.campaigns.filter(user=request.user):
                if campaign.name not in campaign_names:
                    campaign_names.append(campaign.name)
        campaigns = ', '.join(campaign_names)
        
        # Write row
        writer.writerow([
            subscriber.email,
            subscriber.first_name or '',
            subscriber.last_name or '',
            subscriber.full_name or '',
            lists,
            campaigns,
            'Active' if subscriber.is_active else 'Inactive',
            'Yes' if subscriber.confirmed else 'No',
            subscriber.created_at.strftime('%Y-%m-%d %H:%M:%S') if subscriber.created_at else ''
        ])
    
    return response